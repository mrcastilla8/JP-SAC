import io
import openpyxl
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from sgpi_cmee.api.v1.schemas import ExcelGenerationRequest

# Colores institucionales SGPI (Slate-Blue & Teal)
COLOR_TITLE_BG = "1E293B"       # Slate Blue Oscuro (#1E293B)
COLOR_HEADER_BG = "334155"      # Slate Blue Mediano (#334155)
COLOR_ZEBRA_BG = "F8FAFC"       # Gris Azulado Ultra Claro (#F8FAFC)
COLOR_BORDER = "E2E8F0"         # Gris Suave (#E2E8F0)
COLOR_WHITE = "FFFFFF"

def export_to_excel_generic(request: ExcelGenerationRequest) -> io.BytesIO:
    """
    Motor genérico de exportación Excel. 
    Agnóstico al caso de uso: renderiza hojas basándose únicamente en matrices planas (headers y data).
    """
    wb = openpyxl.Workbook()
    
    # Eliminar hoja por defecto al inicio si vamos a crear dinámicamente
    default_sheet = wb.active
    
    font_main_title = Font(name="Segoe UI", size=15, bold=True, color=COLOR_WHITE)
    font_section_header = Font(name="Segoe UI", size=11, bold=True, color=COLOR_WHITE)
    font_data = Font(name="Segoe UI", size=10)
    font_meta = Font(name="Segoe UI", size=9, italic=True)
    
    fill_main_title = PatternFill(start_color=COLOR_TITLE_BG, end_color=COLOR_TITLE_BG, fill_type="solid")
    fill_header = PatternFill(start_color=COLOR_HEADER_BG, end_color=COLOR_HEADER_BG, fill_type="solid")
    fill_zebra = PatternFill(start_color=COLOR_ZEBRA_BG, end_color=COLOR_ZEBRA_BG, fill_type="solid")
    fill_white = PatternFill(start_color=COLOR_WHITE, end_color=COLOR_WHITE, fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color=COLOR_BORDER),
        right=Side(style='thin', color=COLOR_BORDER),
        top=Side(style='thin', color=COLOR_BORDER),
        bottom=Side(style='thin', color=COLOR_BORDER)
    )
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    
    fecha_str = datetime.now().strftime("%d/%m/%Y %H:%M")

    for i, sheet_def in enumerate(request.sheets):
        if i == 0:
            ws = default_sheet
            ws.title = sheet_def.sheet_name
        else:
            ws = wb.create_sheet(title=sheet_def.sheet_name)
            
        ws.views.sheetView[0].showGridLines = True
        
        # Calcular ancho máximo de encabezados para el título principal
        num_cols = 2 # mínimo 2 para la métrica
        if sheet_def.headers:
            num_cols = max(num_cols, len(sheet_def.headers))
        
        # Título Principal
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
        ws.cell(1, 1, sheet_def.title).font = font_main_title
        ws.cell(1, 1).fill = fill_main_title
        ws.cell(1, 1).alignment = align_center
        ws.row_dimensions[1].height = 40
        
        # Metadata
        total_regs = len(sheet_def.data) if sheet_def.data else 0
        meta_text = f"Generado por: {request.usuario_generador}   |   Fecha: {fecha_str}"
        if sheet_def.data is not None:
            meta_text += f"   |   Registros: {total_regs}"
            
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
        ws.cell(2, 1, meta_text).font = font_meta
        ws.cell(2, 1).alignment = align_left
        ws.row_dimensions[2].height = 18

        row_idx = 3

        # Si el reporte es tipo "Métricas" (clave-valor)
        if sheet_def.metrics:
            ws.cell(row_idx, 1, sheet_def.metrics_title).font = font_section_header
            ws.cell(row_idx, 1).fill = fill_header
            ws.cell(row_idx, 1).border = thin_border
            ws.cell(row_idx, 2, sheet_def.metrics_value_title).font = font_section_header
            ws.cell(row_idx, 2).fill = fill_header
            ws.cell(row_idx, 2).border = thin_border
            ws.row_dimensions[row_idx].height = 25
            row_idx += 1
            
            for m, v in sheet_def.metrics:
                row_fill = fill_zebra if row_idx % 2 == 1 else fill_white
                c1 = ws.cell(row=row_idx, column=1, value=m)
                c2 = ws.cell(row=row_idx, column=2, value=v)
                for c in [c1, c2]:
                    c.font = font_data
                    c.fill = row_fill
                    c.border = thin_border
                c1.alignment = align_left
                c2.alignment = align_center
                ws.row_dimensions[row_idx].height = 22
                row_idx += 1

        # Si el reporte es tipo "Tabla" (columnas y datos)
        if sheet_def.headers and sheet_def.data is not None:
            # Si hubo métricas antes, dejamos un espacio
            if sheet_def.metrics:
                row_idx += 1
                
            for col, h in enumerate(sheet_def.headers, 1):
                cell = ws.cell(row=row_idx, column=col, value=h)
                cell.font = font_section_header
                cell.fill = fill_header
                cell.alignment = align_center
                cell.border = thin_border
            ws.row_dimensions[row_idx].height = 25
            row_idx += 1
            
            for data_row in sheet_def.data:
                row_fill = fill_zebra if row_idx % 2 == 1 else fill_white
                for col, val in enumerate(data_row, 1):
                    cell = ws.cell(row=row_idx, column=col, value=val)
                    cell.font = font_data
                    cell.fill = row_fill
                    cell.border = thin_border
                    
                    # Intentar inferir alineación y formato
                    if isinstance(val, (int, float)):
                        if isinstance(val, float):
                            cell.number_format = "#,##0.00"
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    elif val in ["Sí", "No", "N/A"] or (isinstance(val, str) and len(val) < 15 and not " " in val):
                        cell.alignment = align_center
                    else:
                        cell.alignment = align_left
                        
                ws.row_dimensions[row_idx].height = 22
                row_idx += 1

    # Auto-ajustar ancho de las columnas
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col[1:]:
                val = str(cell.value or '')
                if cell.alignment and cell.alignment.wrap_text:
                    max_len = max(max_len, 35)
                else:
                    max_len = max(max_len, len(val))
            sheet.column_dimensions[col_letter].width = max(max_len + 3, 15)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
