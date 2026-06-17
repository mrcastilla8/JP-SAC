import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from cybertesis_connector.core.models import QueryResultsModel


def export_to_excel(results: QueryResultsModel, output_path: str) -> str:
    """
    Exporta la colección de tesis de QueryResultsModel a un archivo de Excel Premium.
    Aplica una paleta corporativa elegante (Slate-Blue & Teal), auto-ajuste de columnas
    y formato tipográfico profesional.
    """
    # Resolver ruta absoluta y asegurar directorio
    abs_path = os.path.abspath(output_path)
    os.makedirs(os.path.dirname(abs_path), exist_ok=True)

    # Crear libro y hoja
    wb = Workbook()
    ws = wb.active
    ws.title = "Tesis Centralizadas"

    # Habilitar líneas de cuadrícula explícitamente
    ws.views.sheetView[0].showGridLines = True

    # 1. Estilos y Paleta de Colores
    font_family = "Segoe UI"

    # Fuentes
    title_font = Font(name=font_family, size=16, bold=True, color="1B365D")  # Dark Navy Blue
    subtitle_font = Font(name=font_family, size=10, italic=True, color="555555")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")  # White text
    data_font = Font(name=font_family, size=10, color="333333")
    link_font = Font(name=font_family, size=10, bold=True, color="008080", underline="single")  # Teal Link

    # Fills
    header_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")  # Slate Dark Blue
    zebra_fill = PatternFill(start_color="F4F7FA", end_color="F4F7FA", fill_type="solid")  # Ice Blue/Grey

    # Borders
    thin_border_side = Side(border_style="thin", color="E0E0E0")
    data_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

    # Alignments
    left_align = Alignment(horizontal="left", vertical="center")
    center_align = Alignment(horizontal="center", vertical="center")

    # 2. Agregar Encabezados de Título del Reporte
    ws.merge_cells("A1:H1")
    ws["A1"] = "REPORTE DE PRODUCCIÓN CIENTÍFICA - CYBERTESIS UNMSM"
    ws["A1"].font = title_font
    ws["A1"].alignment = left_align
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:H2")
    created_date = (
        ws.parent.properties.created.strftime('%d/%m/%Y')
        if ws.parent.properties.created else 'Hoy'
    )
    ws["A2"] = (
        f"Búsqueda realizada: '{results.query}' | "
        f"Tesis recuperadas: {len(results.resultados)} | "
        f"Fecha: {created_date}"
    )
    ws["A2"].font = subtitle_font
    ws["A2"].alignment = left_align
    ws.row_dimensions[2].height = 18

    # Espacio libre
    ws.row_dimensions[3].height = 10

    # 3. Encabezados de la Tabla (Fila 4)
    headers = [
        "N°",
        "Título de la Tesis / Investigación",
        "Tesistas / Autores",
        "Asesores oficiales",
        "Año",
        "Fecha Emisión",
        "Grado Académico",
        "Palabras Clave",
        "Enlace de Cybertesis",
    ]

    start_row = 4
    ws.row_dimensions[start_row].height = 26

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align if col_idx in [1, 5, 6, 7] else left_align
        cell.border = data_border

    # 4. Inserción de Datos (Fila 5 en adelante)
    current_row = 5
    for idx, tesis in enumerate(results.resultados, 1):
        ws.row_dimensions[current_row].height = 22

        # Determinar cebra
        row_fill = zebra_fill if idx % 2 == 0 else None

        # Mapear valores
        values = [
            idx,
            tesis.titulo,
            ", ".join(tesis.autores),
            ", ".join(tesis.asesores) if tesis.asesores else "No especificado en metadatos",
            tesis.anio_publicacion,
            tesis.fecha_sustentacion or "No disponible",
            tesis.grado_academico,
            ", ".join(tesis.palabras_clave) if tesis.palabras_clave else "Sin keywords",
            None,  # Reservado para fórmula hipervínculo
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=current_row, column=col_idx)

            if col_idx == 9:
                # Escribir fórmula de hipervínculo premium en openpyxl
                cell.value = f'=HYPERLINK("{tesis.url_repositorio}", "Ver Tesis")'
                cell.font = link_font
                cell.alignment = center_align
            else:
                cell.value = val
                cell.font = data_font
                cell.alignment = center_align if col_idx in [1, 5, 6, 7] else left_align

            if row_fill:
                cell.fill = row_fill
            cell.border = data_border

        current_row += 1

    # 5. Ajuste automático inteligente del ancho de columnas
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        # No considerar filas de títulos en el cálculo de ancho para evitar columnas sobre-extendidas
        for cell in col:
            if cell.row >= start_row and cell.value:
                # Si es una fórmula de hipervínculo, medir el largo del texto del link
                cell_str = str(cell.value)
                if cell_str.startswith("=HYPERLINK"):
                    cell_str = "Ver Tesis"
                max_len = max(max_len, len(cell_str))

        # Ajustar con margen
        ws.column_dimensions[col_letter].width = max(max_len + 4, 10)

    # El título "Título de la Tesis" o "Resumen" puede requerir un tope de ancho para no ser gigante
    ws.column_dimensions["B"].width = 50  # Título
    ws.column_dimensions["C"].width = 30  # Autores
    ws.column_dimensions["D"].width = 30  # Asesores
    ws.column_dimensions["H"].width = 35  # Palabras Clave
    ws.column_dimensions["I"].width = 16  # Link

    # Guardar archivo
    wb.save(abs_path)
    return abs_path
