import json
from pathlib import Path
from typing import List, Union, Dict, Optional
from pydantic import BaseModel
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from vrip_connector.core.models import ConvocatoriaModel, ProyectoModel, TesisModel

# Visual Identity Palette (Slate-Blue & Teal)
COLOR_TITLE_BG = "1E293B"       # Slate Blue Dark (#1E293B)
COLOR_HEADER_BG = "334155"      # Slate Blue Medium (#334155)
COLOR_ZEBRA_BG = "F8FAFC"       # Light Blue-Gray (#F8FAFC)
COLOR_BORDER = "E2E8F0"         # Light border gray (#E2E8F0)
COLOR_WHITE = "FFFFFF"
COLOR_TEAL_ACCENT = "0D9488"    # Teal Accent (#0D9488)

def export_data(
    data: Union[List[BaseModel], Dict[str, List[BaseModel]]],
    output_path: Optional[str] = None,
    format_type: str = "json",
    quiet: bool = False
):
    """
    Controller for exporting scraped research data to JSON or Premium Excel.
    """
    format_type = format_type.lower()
    
    if format_type == "json":
        # Package and export to JSON
        export_data_json(data, output_path, quiet)
    elif format_type == "excel":
        if not output_path:
            output_path = "reporte_investigaciones_vrip.xlsx"
        export_data_excel(data, output_path, quiet)
    else:
        raise ValueError(f"Formato '{format_type}' no soportado. Elija 'json' o 'excel'.")

def export_data_json(data: Union[List[BaseModel], Dict[str, List[BaseModel]]], output_path: Optional[str], quiet: bool):
    # If it's a dict of lists, serialize each element
    if isinstance(data, dict):
        serialized_dict = {}
        for key, lst in data.items():
            serialized_dict[key] = [item.model_dump() for item in lst]
        json_data = json.dumps(serialized_dict, indent=2, ensure_ascii=False)
    else:
        json_data = json.dumps([item.model_dump() for item in data], indent=2, ensure_ascii=False)
        
    if output_path:
        out_file = Path(output_path)
        out_file.parent.mkdir(parents=True, exist_ok=True)
        with open(out_file, "w", encoding="utf-8") as f:
            f.write(json_data)
        if not quiet:
            print(f"Éxito: Archivo JSON guardado en: {out_file.resolve()}")
    else:
        # Output to stdout
        print(json_data)

def export_data_excel(data: Union[List[BaseModel], Dict[str, List[BaseModel]]], path: str, quiet: bool):
    """Generates a professional corporate corporate quality Excel workbook using openpyxl."""
    wb = openpyxl.Workbook()
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Standard styles definition
    font_main_title = Font(name="Segoe UI", size=14, bold=True, color=COLOR_WHITE)
    font_section_header = Font(name="Segoe UI", size=10, bold=True, color=COLOR_WHITE)
    font_data = Font(name="Segoe UI", size=9)
    font_data_bold = Font(name="Segoe UI", size=9, bold=True)
    font_link = Font(name="Segoe UI", size=9, color="0000FF", underline="single")
    
    fill_main_title = PatternFill(start_color=COLOR_TITLE_BG, end_color=COLOR_TITLE_BG, fill_type="solid")
    fill_header = PatternFill(start_color=COLOR_HEADER_BG, end_color=COLOR_HEADER_BG, fill_type="solid")
    fill_zebra = PatternFill(start_color=COLOR_ZEBRA_BG, end_color=COLOR_ZEBRA_BG, fill_type="solid")
    fill_white = PatternFill(start_color=COLOR_WHITE, end_color=COLOR_WHITE, fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color=COLOR_BORDER),
        right=Side(style='thin', color=COLOR_BORDER),
        top=Side(style='thin', color=COLOR_BORDER),
        bottom=Side(style='thin', color=COLOR_BORDER)
    )
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_wrap_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Normalize data into dict of lists
    sheets_data: Dict[str, List[BaseModel]] = {}
    if isinstance(data, dict):
        sheets_data = data
    else:
        # If it's a list, check type of first item
        if data:
            first_item = data[0]
            if isinstance(first_item, ConvocatoriaModel):
                sheets_data["Convocatorias"] = data
            elif isinstance(first_item, ProyectoModel):
                sheets_data["Proyectos_Resoluciones"] = data
            elif isinstance(first_item, TesisModel):
                sheets_data["Cybertesis_Tesis"] = data
        else:
            # Empty list
            sheets_data["Resultados"] = []

    # Write each sheet
    for sheet_name, item_list in sheets_data.items():
        ws = wb.create_sheet(title=sheet_name)
        ws.views.sheetView[0].showGridLines = True
        
        if not item_list:
            ws.cell(row=1, column=1, value="No se encontraron registros.").font = font_data
            continue
            
        first_item = item_list[0]
        
        # Structure headers and layout based on model type
        if isinstance(first_item, ConvocatoriaModel):
            title_text = "CONVOCATORIAS DE FINANCIAMIENTO VIGENTES - PORTAL VRIP"
            headers = ["Título de Convocatoria / Programa", "Entidad Promotora", "Publicación", "Cierre ISO", "Cierre Original", "Días Restantes", "Enlace Bases / Directivas"]
            
            # Setup title
            ws.merge_cells("A1:G1")
            ws["A1"] = title_text
            ws["A1"].font = font_main_title
            ws["A1"].fill = fill_main_title
            ws["A1"].alignment = align_center
            ws.row_dimensions[1].height = 36
            
            # Setup headers
            for col_idx, h in enumerate(headers, 1):
                cell = ws.cell(row=2, column=col_idx, value=h)
                cell.font = font_section_header
                cell.fill = fill_header
                cell.alignment = align_center
                cell.border = thin_border
            ws.row_dimensions[2].height = 24
            
            # Fill rows
            row_idx = 3
            for item in item_list:
                row_fill = fill_zebra if row_idx % 2 == 1 else fill_white
                
                c1 = ws.cell(row=row_idx, column=1, value=item.titulo)
                c2 = ws.cell(row=row_idx, column=2, value=item.entidad_promotora)
                c3 = ws.cell(row=row_idx, column=3, value=item.fecha_publicacion or "No indicada")
                c4 = ws.cell(row=row_idx, column=4, value=item.plazo_cierre or "N/A")
                c5 = ws.cell(row=row_idx, column=5, value=item.plazo_cierre_original)
                c6 = ws.cell(row=row_idx, column=6, value=item.dias_restantes if item.dias_restantes is not None else "-")
                c7 = ws.cell(row=row_idx, column=7, value="Ver bases / link")
                c7.hyperlink = item.enlace
                c7.font = font_link
                
                for c in [c1, c2, c3, c4, c5, c6, c7]:
                    if c != c7:
                        c.font = font_data
                    c.fill = row_fill
                    c.border = thin_border
                    
                c1.alignment = align_wrap_left
                c2.alignment = align_left
                c3.alignment = align_center
                c4.alignment = align_center
                c5.alignment = align_center
                c6.alignment = align_center
                c7.alignment = align_center
                
                ws.row_dimensions[row_idx].height = 20
                row_idx += 1

        elif isinstance(first_item, ProyectoModel):
            title_text = "RESOLUCIONES RECTORALES Y PROYECTOS APROBADOS - VRIP"
            headers = ["Código", "Programa", "Título de Proyecto / Resolución", "Responsable", "Facultad", "Presupuesto", "Nro Resolución", "Aprobación", "Año", "Enlace Post"]
            
            ws.merge_cells("A1:J1")
            ws["A1"] = title_text
            ws["A1"].font = font_main_title
            ws["A1"].fill = fill_main_title
            ws["A1"].alignment = align_center
            ws.row_dimensions[1].height = 36
            
            for col_idx, h in enumerate(headers, 1):
                cell = ws.cell(row=2, column=col_idx, value=h)
                cell.font = font_section_header
                cell.fill = fill_header
                cell.alignment = align_center
                cell.border = thin_border
            ws.row_dimensions[2].height = 24
            
            row_idx = 3
            for item in item_list:
                row_fill = fill_zebra if row_idx % 2 == 1 else fill_white
                
                c1 = ws.cell(row=row_idx, column=1, value=item.codigo_proyecto or "-")
                c2 = ws.cell(row=row_idx, column=2, value=item.codigo_programa)
                c3 = ws.cell(row=row_idx, column=3, value=item.titulo)
                c4 = ws.cell(row=row_idx, column=4, value=item.responsable)
                c5 = ws.cell(row=row_idx, column=5, value=item.facultad)
                c6 = ws.cell(row=row_idx, column=6, value=item.monto_financiado if item.monto_financiado is not None else "-")
                c7 = ws.cell(row=row_idx, column=7, value=item.numero_resolucion or "Por registrar")
                c8 = ws.cell(row=row_idx, column=8, value=item.fecha_aprobacion or "-")
                c9 = ws.cell(row=row_idx, column=9, value=item.anio_academico)
                c10 = ws.cell(row=row_idx, column=10, value="Ver Portal")
                c10.hyperlink = item.enlace_vrip
                c10.font = font_link
                
                for c in [c1, c2, c3, c4, c5, c6, c7, c8, c9, c10]:
                    if c != c10:
                        c.font = font_data
                    c.fill = row_fill
                    c.border = thin_border
                    
                c1.alignment = align_center
                c2.alignment = align_center
                c3.alignment = align_wrap_left
                c4.alignment = align_left
                c5.alignment = align_center
                c6.alignment = align_right
                if item.monto_financiado is not None:
                    c6.number_format = "S/. #,##0.00"
                c7.alignment = align_center
                c8.alignment = align_center
                c9.alignment = align_center
                c10.alignment = align_center
                
                ws.row_dimensions[row_idx].height = 22
                row_idx += 1

        elif isinstance(first_item, TesisModel):
            title_text = "TESIS Y PRODUCCIÓN ACADÉMICA RECUPERADA - CYBERTESIS"
            headers = ["Título de la Tesis", "Autores / Tesistas", "Año Publicación", "Enlace de Repositorio (Handle)"]
            
            ws.merge_cells("A1:D1")
            ws["A1"] = title_text
            ws["A1"].font = font_main_title
            ws["A1"].fill = fill_main_title
            ws["A1"].alignment = align_center
            ws.row_dimensions[1].height = 36
            
            for col_idx, h in enumerate(headers, 1):
                cell = ws.cell(row=2, column=col_idx, value=h)
                cell.font = font_section_header
                cell.fill = fill_header
                cell.alignment = align_center
                cell.border = thin_border
            ws.row_dimensions[2].height = 24
            
            row_idx = 3
            for item in item_list:
                row_fill = fill_zebra if row_idx % 2 == 1 else fill_white
                
                c1 = ws.cell(row=row_idx, column=1, value=item.titulo)
                c2 = ws.cell(row=row_idx, column=2, value=item.autores)
                c3 = ws.cell(row=row_idx, column=3, value=item.anio_publicacion)
                c4 = ws.cell(row=row_idx, column=4, value=item.enlace_handle)
                c4.hyperlink = item.enlace_handle
                c4.font = font_link
                
                for c in [c1, c2, c3, c4]:
                    if c != c4:
                        c.font = font_data
                    c.fill = row_fill
                    c.border = thin_border
                    
                c1.alignment = align_wrap_left
                c2.alignment = align_left
                c3.alignment = align_center
                c4.alignment = align_center
                
                ws.row_dimensions[row_idx].height = 20
                row_idx += 1

        # Auto-adjust column widths based on contents
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            
            # Skip the main title in row 1 for width calculation
            for cell in col[1:]:
                val = str(cell.value or '')
                if cell.alignment.wrap_text:
                    max_len = max(max_len, 35)  # Cap wrapped fields to a standard width
                else:
                    max_len = max(max_len, len(val))
                    
            # Apply padding
            ws.column_dimensions[col_letter].width = max(max_len + 3, 11)

    # Save to disk
    output_file = Path(path)
    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_file)
    
    if not quiet:
        print(f"Éxito: Reporte de Excel Premium guardado en: {output_file.resolve()}")
